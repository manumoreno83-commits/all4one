"""
Excel (.xlsx) operations using openpyxl.

Workflow: download from Drive → edit in memory → upload back to Drive.
"""

import io
import logging

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string

logger = logging.getLogger("mcp_gdrive.excel")
audit = logging.getLogger("mcp_gdrive.audit")


def _parse_range(range_str: str) -> tuple[int, int, int, int]:
    """
    Parse an A1-style range like 'A1:C10' into (min_row, min_col, max_row, max_col).
    Also supports single cell like 'B5'.
    """
    from openpyxl.utils.cell import range_boundaries
    return range_boundaries(range_str)


def read_range(
    file_data: bytes,
    sheet_name: str | None = None,
    range_str: str | None = None,
) -> dict:
    """
    Read values from an Excel file in memory.
    If sheet_name is None, uses the active sheet.
    If range_str is None, reads all data.
    """
    wb = load_workbook(io.BytesIO(file_data), read_only=True, data_only=True)

    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise ValueError(
                f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"
            )
        ws = wb[sheet_name]
    else:
        ws = wb.active
        sheet_name = ws.title

    if range_str:
        min_col, min_row, max_col, max_row = _parse_range(range_str)
        rows = []
        for row in ws.iter_rows(
            min_row=min_row,
            max_row=max_row,
            min_col=min_col,
            max_col=max_col,
            values_only=True,
        ):
            rows.append([_cell_to_str(c) for c in row])
    else:
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append([_cell_to_str(c) for c in row])

    wb.close()
    logger.info("Read %d rows from Excel sheet '%s'", len(rows), sheet_name)

    return {
        "sheet_name": sheet_name,
        "range": range_str or "all",
        "values": rows,
        "row_count": len(rows),
    }


def write_range(
    file_data: bytes,
    sheet_name: str | None,
    range_str: str,
    values: list[list],
    dry_run: bool = False,
) -> bytes:
    """
    Write values into an Excel file in memory.
    Returns the modified file as bytes.
    """
    if not values:
        raise ValueError("values cannot be empty")

    if dry_run:
        logger.info("[DRY RUN] Would write %d rows to Excel '%s' %s", len(values), sheet_name, range_str)
        return file_data  # unchanged

    wb = load_workbook(io.BytesIO(file_data))

    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise ValueError(
                f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"
            )
        ws = wb[sheet_name]
    else:
        ws = wb.active
        sheet_name = ws.title

    min_col, min_row, _, _ = _parse_range(range_str)

    cells_written = 0
    for r_offset, row_data in enumerate(values):
        for c_offset, value in enumerate(row_data):
            ws.cell(
                row=min_row + r_offset,
                column=min_col + c_offset,
                value=value,
            )
            cells_written += 1

    audit.info(
        "EXCEL_WRITE | sheet=%s | range=%s | cells=%d",
        sheet_name,
        range_str,
        cells_written,
    )
    logger.info("Wrote %d cells to Excel sheet '%s' %s", cells_written, sheet_name, range_str)

    buffer = io.BytesIO()
    wb.save(buffer)
    wb.close()
    buffer.seek(0)
    return buffer.read()


def _cell_to_str(value) -> str | None:
    """Convert a cell value to a JSON-safe string representation."""
    if value is None:
        return None
    return str(value)
